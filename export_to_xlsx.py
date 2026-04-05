#!/usr/bin/env python3
"""
YRCARKIT Battery Cycle Data -> XLSX Exporter
Scans the most recent session from w_lxdzdb, extracts the last complete
discharge cycle per channel, and writes results to an Excel spreadsheet.

Auto-detects active channels from the session data.
Each run tests N modules and maps them to cell IDs 1-28 in a battery sheet.
"""

import os
import re
import sys
import sqlite3
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("\nMissing openpyxl. Run:  pip install openpyxl\n")
    sys.exit(1)

# -- CONFIG -------------------------------------------------------
DB_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "w_lxdzdb")
XLSX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "battery_data.xlsx")
MIN_CYCLES = 3   # warn if fewer than this many charge-discharge pairs
# -----------------------------------------------------------------

METRICS = [
    "Discharge Capacity (Ah)",
    "Internal Resistance (mOhm)",
    "End Voltage (V)",
    "Duration (min)",
    "Test Date",
    "Session ID",
    "Channel Used",
]


# -- DB SCANNING --------------------------------------------------

def scan_sessions(folder):
    """Group all DB files by session. Returns {session_key: {ch: filepath}}."""
    dbs = sorted(Path(folder).glob("A*_CH*_04.db"))
    sessions = {}

    for fp in dbs:
        m = re.match(r"A(\d{8})(\d{6})_CH(\d+)_04\.db", fp.name)
        if not m:
            continue
        date_str, time_str, ch = m.group(1), m.group(2), int(m.group(3))

        key = f"{date_str}_{time_str[:4]}"
        ts = datetime.strptime(f"{date_str} {time_str}", "%Y%m%d %H%M%S")

        if key not in sessions:
            sessions[key] = {"key": key, "timestamp": ts,
                             "date": ts.strftime("%Y-%m-%d"), "channels": {}}
        sessions[key]["channels"][ch] = str(fp)

    return sessions


def read_cycles(filepath):
    """Read all cycle tables from a channel DB. Returns sorted list."""
    conn = sqlite3.connect(filepath)
    cur = conn.cursor()
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
    tables = [r[0] for r in cur.fetchall()]

    cycles = []
    for t in tables:
        tm = re.match(r"([CF])(\d+)_CH", t)
        if not tm:
            continue
        kind = tm.group(1)          # C = charge, F = discharge
        seq = int(tm.group(2))

        cur.execute(f'SELECT MAX(cap), MAX(tim), COUNT(*) FROM "{t}"')
        max_cap, max_tim, cnt = cur.fetchone()

        cur.execute(f'SELECT vol FROM "{t}" ORDER BY id DESC LIMIT 1')
        r = cur.fetchone()
        v_end = round(r[0], 3) if r else None

        # IR: voltage drop when current kicks in
        ir_mohm = None
        cur.execute(f'SELECT vol, cur FROM "{t}" ORDER BY id ASC LIMIT 15')
        rows = cur.fetchall()
        v_rest = v_load = i_load = None
        for v, c in rows:
            if c == 0.0 and v_rest is None:
                v_rest = v
            elif c >= 1.0 and v_rest is not None and v_load is None:
                v_load = v
                i_load = c
                break
        if v_rest and v_load and i_load:
            ir = round(abs(v_rest - v_load) / i_load * 1000, 1)
            if 2 <= ir <= 200:
                ir_mohm = ir

        cycles.append({
            "table": t, "kind": kind, "seq": seq,
            "cap_ah": round(max_cap, 3) if max_cap else 0,
            "dur_s": max_tim or 0, "v_end": v_end,
            "rows": cnt, "ir_mohm": ir_mohm,
        })

    conn.close()
    cycles.sort(key=lambda x: x["seq"])
    return cycles


def find_target_discharge(cycles):
    """
    Pick the last complete discharge:
      - If last table is Charge  -> 2nd-to-last  (should be Discharge)
      - If last table is Discharge -> 3rd-to-last (should be Discharge)
    """
    if len(cycles) < 2:
        return None
    last = cycles[-1]
    if last["kind"] == "C":
        target = cycles[-2]
    else:
        if len(cycles) < 3:
            return None
        target = cycles[-3]
    return target if target["kind"] == "F" else None


def count_discharge_cycles(cycles):
    """Count total discharge cycles (F tables)."""
    return sum(1 for c in cycles if c["kind"] == "F")


# -- XLSX ---------------------------------------------------------

def get_or_create_workbook(path):
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def get_or_create_sheet(wb, name):
    name = name.upper()
    if name in wb.sheetnames:
        return wb[name]

    ws = wb.create_sheet(title=name)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    center = Alignment(horizontal="center")

    # Column A: metric labels
    ws.cell(row=1, column=1, value="Metric").font = hdr_font
    ws.cell(row=1, column=1).fill = hdr_fill
    ws.column_dimensions["A"].width = 30

    # Columns B-AC: Cell 1-28
    for i in range(1, 29):
        c = ws.cell(row=1, column=i + 1, value=f"Cell {i}")
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        ws.column_dimensions[c.column_letter].width = 15

    # Metric row labels
    for idx, label in enumerate(METRICS, start=2):
        ws.cell(row=idx, column=1, value=label).font = Font(bold=True)

    return ws


def write_cell(ws, cell_id, data):
    col = cell_id + 1       # col B = Cell 1
    center = Alignment(horizontal="center")
    ws.cell(row=2, column=col, value=data["cap_ah"]).alignment = center
    ws.cell(row=3, column=col, value=data["ir_mohm"]).alignment = center
    ws.cell(row=4, column=col, value=data["v_end"]).alignment = center
    ws.cell(row=5, column=col, value=round(data["dur_s"] / 60, 1)).alignment = center
    ws.cell(row=6, column=col, value=data["date"]).alignment = center
    ws.cell(row=7, column=col, value=data["session"]).alignment = center
    ws.cell(row=8, column=col, value=f"CH{data['ch']}").alignment = center


# -- MAIN ---------------------------------------------------------

def main():
    print("=" * 55)
    print("  YRCARKIT  ->  XLSX  Battery Data Exporter")
    print("=" * 55)

    # 1. Scan sessions
    print(f"\nScanning: {DB_FOLDER}")
    sessions = scan_sessions(DB_FOLDER)
    if not sessions:
        print("No sessions found!"); return

    latest_key = max(sessions.keys())
    latest = sessions[latest_key]
    all_channels = sorted(latest["channels"].keys())

    print(f"\nMost recent session: {latest_key}  ({latest['date']})")
    print(f"  Channels found: {all_channels}")

    # Ask which channels to skip
    skip_input = input("\nChannels to skip (e.g. 3 or 3,5 or press Enter for none): ").strip()
    skip_channels = set()
    if skip_input:
        for s in re.split(r"[,\s]+", skip_input):
            if s.isdigit():
                skip_channels.add(int(s))

    active_channels = [ch for ch in all_channels if ch not in skip_channels]
    if skip_channels:
        print(f"  Skipping: {sorted(skip_channels)}")
    print(f"  Active channels: {active_channels}  ({len(active_channels)} total)")

    # 2. Analyze each channel
    print("\nCycle analysis:")
    results = {}
    all_ok = True

    for ch in active_channels:
        cycles = read_cycles(latest["channels"][ch])
        num_dis = count_discharge_cycles(cycles)
        target = find_target_discharge(cycles)

        tag = "OK" if num_dis >= MIN_CYCLES and target else "WARN"
        if tag == "WARN":
            all_ok = False

        cap_str = f"{target['cap_ah']}Ah" if target else "--"
        ir_str = f"{target['ir_mohm']}mOhm" if target and target["ir_mohm"] else "--"
        print(f"  CH{ch}: {num_dis} discharges  |  target={target['table'] if target else 'N/A'}  "
              f"|  cap={cap_str}  IR={ir_str}  [{tag}]")

        if target:
            results[ch] = {
                "cap_ah": target["cap_ah"],
                "ir_mohm": target["ir_mohm"],
                "v_end": target["v_end"],
                "dur_s": target["dur_s"],
                "date": latest["date"],
                "session": latest_key,
                "ch": ch,
            }

    if not results:
        print("\nNo valid discharge data found!"); return

    if not all_ok:
        ans = input(f"\nSome channels have fewer than {MIN_CYCLES} discharge cycles. Continue anyway? (y/n): ").strip().lower()
        if ans != "y":
            print("Aborted."); return

    # 3. Ask battery ID
    print()
    battery = input("What battery is this? (A/B/C/D/...): ").strip().upper()
    if not battery or not battery.isalpha():
        print("Invalid. Use a letter (A-Z)."); return

    # 4. Ask cell range
    num_active = len(active_channels)
    cell_input = input(f"What cell IDs did you charge? ({num_active} cells, e.g. 1-{num_active}): ").strip()
    m = re.match(r"(\d+)\s*[-\u2013]\s*(\d+)", cell_input)
    if not m:
        print(f"Invalid range. Use format like: 1-{num_active}"); return

    start_id = int(m.group(1))
    end_id = int(m.group(2))
    cell_ids = list(range(start_id, end_id + 1))

    if len(cell_ids) != num_active:
        print(f"Need exactly {num_active} cells (matching {num_active} active channels), got {len(cell_ids)}."); return
    if start_id < 1 or end_id > 28:
        print("Cell IDs must be 1-28."); return

    # 5. Map channels -> cell IDs and confirm
    print(f"\nMapping (Battery {battery}):")
    mapping = {}
    for ch, cid in zip(active_channels, cell_ids):
        if ch in results:
            mapping[cid] = results[ch]
            ir_val = results[ch]['ir_mohm'] if results[ch]['ir_mohm'] else '--'
            print(f"  CH{ch}  ->  Cell {cid:2d}  |  {results[ch]['cap_ah']}Ah  "
                  f"IR={ir_val}mOhm  Vend={results[ch]['v_end']}V")

    confirm = input(f"\nWrite {len(mapping)} cells to sheet '{battery}'? (y/n): ").strip().lower()
    if confirm != "y":
        print("Aborted."); return

    # 6. Write XLSX
    wb = get_or_create_workbook(XLSX_PATH)
    ws = get_or_create_sheet(wb, battery)

    for cid, data in mapping.items():
        write_cell(ws, cid, data)

    wb.save(XLSX_PATH)
    print(f"\n{'=' * 55}")
    print(f"  SAVED: {XLSX_PATH}")
    print(f"  Sheet: {battery}   Cells: {start_id}-{end_id}")
    print(f"{'=' * 55}")


if __name__ == "__main__":
    main()
